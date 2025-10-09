from pathlib import Path
from typing import List
from openpyxl import load_workbook
from core.utils import Chunk, estimate_tokens

def parse_xlsx(file_path: Path, max_tokens: int, overlap: int) -> List[Chunk]:
    chunks = []
    
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            # Extract headers and data
            rows_data = []
            for row in sheet.iter_rows(values_only=True):
                row_text = ' | '.join(str(cell) if cell is not None else '' for cell in row)
                if row_text.strip():
                    rows_data.append(row_text)
            
            if rows_data:
                sheet_text = '\n'.join(rows_data)
                token_count = estimate_tokens(sheet_text)
                
                # If sheet is too large, split it
                if token_count > max_tokens:
                    from core.utils import split_into_chunks
                    sub_chunks = split_into_chunks(sheet_text, max_tokens, overlap)
                    for i, sub_text in enumerate(sub_chunks):
                        chunks.append(Chunk(
                            text=sub_text,
                            section_path=f"sheet_{sheet_name}_part_{i}",
                            token_count=estimate_tokens(sub_text),
                            extra_meta={"sheet": sheet_name}
                        ))
                else:
                    chunks.append(Chunk(
                        text=sheet_text,
                        section_path=f"sheet_{sheet_name}",
                        token_count=token_count,
                        extra_meta={"sheet": sheet_name}
                    ))
    except Exception as e:
        print(f"Error parsing XLSX {file_path}: {e}")
    
    return chunks